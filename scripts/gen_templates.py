"""
生成 Excel 模板 - 仿 API 框架 templates/*.xlsx
路径: scripts/gen_templates.py
生成：
    templates/test_result_template.xlsx   (UI 用例结果模板，9 列)
    templates/bug_template.xlsx           (Bug 清单模板，6 列)
⚠️ 修改模板列数必须同步修改 utils/excel_export.py 中的 RESULT_HEADERS / BUG_HEADERS。
"""
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment


def make(path, headers, sample_rows=None):
    wb = Workbook()
    ws = wb.active
    ws.title = "模板"
    ws.append(headers)
    head_fill = PatternFill("solid", fgColor="305496")
    head_font = Font(color="FFFFFF", bold=True)
    for i in range(1, len(headers) + 1):
        c = ws.cell(row=1, column=i)
        c.fill = head_fill
        c.font = head_font
        c.alignment = Alignment(horizontal="center", vertical="center")
    if sample_rows:
        for r in sample_rows:
            ws.append(r)
    # 列宽
    for i, h in enumerate(headers, 1):
        ws.column_dimensions[chr(64 + i) if i <= 26 else "A"].width = max(len(h) * 2 + 6, 14)
    wb.save(path)
    print(f"已生成: {path}")


if __name__ == "__main__":
    import os
    os.makedirs("templates", exist_ok=True)
    make("templates/test_result_template.xlsx",
         ["模块", "用例名", "优先级", "状态", "步骤", "错误分类", "错误消息", "耗时(ms)", "截图"],
         [["data_integration", "test_source_list_visible", "P0", "PASS", "打开列表页", "", "", "1234", "见 Allure"]])
    make("templates/bug_template.xlsx",
         ["模块", "Bug标题（可直接导入TAPD）", "错误分类", "错误消息", "用例名", "失败步骤"],
         [["data_integration", "【data_integration】test_source_detail UI → ElementNotFound 元素不可见", "ElementNotFound", "元素不可见", "test_source_detail", ""]])
