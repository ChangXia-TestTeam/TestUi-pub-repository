"""
Excel 导出 - 仿 API 框架（按模板填充，PASS 绿底 / FAIL 红底加粗）
路径: utils/excel_export.py
模板：
    templates/test_result_template.xlsx   → 全量用例结果
    templates/bug_template.xlsx           → 仅 FAIL 的缺陷清单
输出到 UI_output_files/test_ui_results 与 UI_output_files/bug_list。
Bug 标题模板：【{模块}】{用例名} UI {METHOD} → {错误分类} {错误消息}
"""
from __future__ import annotations

import re
from pathlib import Path
from datetime import datetime

from utils import config

# 颜色
PASS_FILL = "C6EFCE"   # 淡绿
FAIL_FILL = "FFC7CE"   # 淡红
FAIL_FONT_COLOR = "9C0006"

# 用例结果表头（与 API 模板保持一致风格，UI 适配）
RESULT_HEADERS = ["模块", "用例名", "优先级", "状态", "步骤", "错误分类", "错误消息", "耗时(ms)", "截图"]
BUG_HEADERS = ["模块", "Bug标题（可直接导入TAPD）", "错误分类", "错误消息", "用例名", "失败步骤"]


def _new_workbook(headers: list[str], rows: list[list]):
    """openpyxl 新建 workbook + 表头 + 数据 + 着色。"""
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "结果"
    ws.append(headers)
    # 表头样式
    head_fill = PatternFill("solid", fgColor="305496")
    head_font = Font(color="FFFFFF", bold=True)
    for col_idx, _ in enumerate(headers, 1):
        c = ws.cell(row=1, column=col_idx)
        c.fill = head_fill
        c.font = head_font
        c.alignment = Alignment(horizontal="center", vertical="center")
    # 数据
    for r in rows:
        ws.append(r)
        status_idx = headers.index("状态") + 1 if "状态" in headers else None
        if status_idx:
            status = r[status_idx - 1]
            color = PASS_FILL if str(status).upper() in ("PASS", "PASSED") else FAIL_FILL
            is_fail = str(status).upper() in ("FAIL", "FAILED", "BROKEN")
            for col_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=ws.max_row, column=col_idx)
                cell.fill = PatternFill("solid", fgColor=color)
                if is_fail:
                    cell.font = Font(color=FAIL_FONT_COLOR, bold=True)
    # 列宽
    for col_idx, h in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = max(len(str(h)) * 2 + 6, 12)
    return wb


def _parse_module(full_name: str) -> str:
    """从 fullName 提取模块名（tests/test_xxx.py::TestClass::test_yyy → xxx）。"""
    m = re.search(r"test_([a-z_]+)\.py", full_name or "")
    return m.group(1) if m else "unknown"


def _parse_priority(full_name: str) -> str:
    return "P0" if "P0" in full_name else ("P1" if "P1" in full_name else "P2")


def _classify_error(msg: str) -> str:
    """简单错误分类（便于 TAPD）。"""
    if not msg:
        return "Unknown"
    s = str(msg).lower()
    if "timeout" in s or "超时" in s:
        return "Timeout"
    if "not visible" in s or "不可见" in s or "not found" in s:
        return "ElementNotFound"
    if "assert" in s:
        return "AssertionFailed"
    if "navigation" in s or "url" in s:
        return "NavigationError"
    return "Unknown"


def export_results(stats: dict) -> dict:
    """导出全量结果 + bug 清单，返回路径。"""
    results_dir = ROOT_child("UI_output_files/test_ui_results")
    bug_dir = ROOT_child("UI_output_files/bug_list")
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")

    # 1) 全量结果
    result_rows = []
    for c in stats.get("failed_cases", []):
        full = c.get("full", "")
        msg = c.get("status_message", "")
        result_rows.append([
            _parse_module(full),
            c.get("name", ""),
            _parse_priority(full),
            "FAIL",
            "",
            _classify_error(msg),
            msg,
            "",
            "见 Allure 附件",
        ])
    # 补充 passed/summary 行
    if not result_rows:
        result_rows.append([config.project_name(), "全部通过", "P0", "PASS", "", "", "", "", ""])
    wb = _new_workbook(RESULT_HEADERS, result_rows)
    results_path = results_dir / f"test_ui_results_{ts}.xlsx"
    wb.save(str(results_path))

    # 2) Bug 清单（仅 FAIL）
    bug_rows = []
    for c in stats.get("failed_cases", []):
        full = c.get("full", "")
        msg = c.get("status_message", "")
        module = _parse_module(full)
        cat = _classify_error(msg)
        title = f"【{module}】{c.get('name', '')} UI → {cat} {msg[:80]}"
        bug_rows.append([module, title, cat, msg, c.get("name", ""), ""])
    if not bug_rows:
        bug_rows.append([config.project_name(), "无失败用例", "", "", "", ""])
    wb2 = _new_workbook(BUG_HEADERS, bug_rows)
    bug_path = bug_dir / f"bug_list_{ts}.xlsx"
    wb2.save(str(bug_path))

    return {"results": str(results_path), "bug_list": str(bug_path)}


def ROOT_child(rel: str) -> Path:
    p = config.ROOT_DIR / rel
    p.mkdir(parents=True, exist_ok=True)
    return p
